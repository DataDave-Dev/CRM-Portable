from datetime import datetime

from app.repositories.reporte_repository import ReporteRepository
from app.utils.logger import AppLogger
from app.utils.db_retry import sanitize_error_message

logger = AppLogger.get_logger(__name__)

# Definición de cada reporte: (nombre legible, columnas a mostrar, alias de cabecera)
REPORTES = {
    "pipeline": {
        "titulo": "Pipeline de Ventas",
        "columnas": [
            "NombreOportunidad", "Empresa", "Contacto", "Etapa",
            "MontoEstimado", "ProbabilidadCierre", "ValorPonderado",
            "FechaCierreEstimada", "Vendedor", "DiasEnPipeline",
        ],
        "cabeceras": [
            "Oportunidad", "Empresa", "Contacto", "Etapa",
            "Monto Estimado", "Prob. %", "Valor Ponderado",
            "Cierre Estimado", "Vendedor", "Días en Pipeline",
        ],
    },
    "vendedores": {
        "titulo": "Rendimiento de Vendedores",
        "columnas": [
            "Vendedor", "OportunidadesAbiertas", "OportunidadesGanadas",
            "OportunidadesPerdidas", "MontoGanado", "MontoPendiente",
            "TasaConversion", "PromedioDiasCierre",
        ],
        "cabeceras": [
            "Vendedor", "Abiertas", "Ganadas", "Perdidas",
            "Monto Ganado", "Monto Pendiente", "Tasa Conv. %", "Días Prom. Cierre",
        ],
    },
    "etapas": {
        "titulo": "Conversión por Etapa",
        "columnas": [
            "Etapa", "TotalOportunidades", "MontoTotal",
            "TicketPromedio", "ProbabilidadEtapa",
        ],
        "cabeceras": [
            "Etapa", "Total Oportunidades", "Monto Total",
            "Ticket Promedio", "Probabilidad %",
        ],
    },
    "campanas": {
        "titulo": "Análisis de Campañas",
        "columnas": [
            "Nombre", "Tipo", "Estado", "FechaEnvio",
            "TotalDestinatarios", "TotalEnviados", "TotalAbiertos", "TotalClics",
            "TasaEntrega", "TasaApertura", "TasaClics", "Propietario",
        ],
        "cabeceras": [
            "Campaña", "Tipo", "Estado", "Fecha Envío",
            "Destinatarios", "Enviados", "Abiertos", "Clics",
            "Entrega %", "Apertura %", "Clics %", "Propietario",
        ],
    },
    "actividad": {
        "titulo": "Actividad de Contactos",
        "columnas": [
            "NombreContacto", "Empresa", "TotalActividades",
            "UltimaActividad", "DiasSinContacto",
            "TotalLlamadas", "TotalReuniones", "TotalCorreos",
        ],
        "cabeceras": [
            "Contacto", "Empresa", "Total Actividades",
            "Última Actividad", "Días sin Contacto",
            "Llamadas", "Reuniones", "Correos",
        ],
    },
}


class ReporteService:

    def __init__(self):
        self._repo = ReporteRepository()

    # ------------------------------------------------------------------ #
    # Obtener datos                                                        #
    # ------------------------------------------------------------------ #

    def obtener_pipeline_ventas(self, fecha_desde=None, fecha_hasta=None):
        try:
            datos = self._repo.get_pipeline_ventas(fecha_desde, fecha_hasta)
            return datos, None
        except Exception as e:
            AppLogger.log_exception(logger, "Error al obtener pipeline de ventas")
            return None, sanitize_error_message(e)

    def obtener_rendimiento_vendedores(self):
        try:
            datos = self._repo.get_rendimiento_vendedores()
            return datos, None
        except Exception as e:
            AppLogger.log_exception(logger, "Error al obtener rendimiento vendedores")
            return None, sanitize_error_message(e)

    def obtener_conversion_etapas(self):
        try:
            datos = self._repo.get_conversion_por_etapa()
            return datos, None
        except Exception as e:
            AppLogger.log_exception(logger, "Error al obtener conversión por etapa")
            return None, sanitize_error_message(e)

    def obtener_analisis_campanas(self, fecha_desde=None, fecha_hasta=None):
        try:
            datos = self._repo.get_analisis_campanas(fecha_desde, fecha_hasta)
            return datos, None
        except Exception as e:
            AppLogger.log_exception(logger, "Error al obtener análisis de campañas")
            return None, sanitize_error_message(e)

    def obtener_actividad_contactos(self, fecha_desde=None, fecha_hasta=None):
        try:
            datos = self._repo.get_actividad_contactos(fecha_desde, fecha_hasta)
            return datos, None
        except Exception as e:
            AppLogger.log_exception(logger, "Error al obtener actividad de contactos")
            return None, sanitize_error_message(e)

    # ------------------------------------------------------------------ #
    # Exportar a Excel                                                     #
    # ------------------------------------------------------------------ #

    def exportar_excel(self, clave_reporte, datos, ruta):
        """
        Exporta los datos de un reporte a un archivo Excel (.xlsx).

        Args:
            clave_reporte: clave del dict REPORTES ('pipeline', 'vendedores', etc.)
            datos: lista de dicts con los datos a exportar
            ruta: ruta completa del archivo destino

        Returns:
            (True, None) si éxito, (None, mensaje_error) si falla
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                PatternFill, Font, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter

            cfg = REPORTES[clave_reporte]
            columnas = cfg["columnas"]
            cabeceras = cfg["cabeceras"]
            titulo = cfg["titulo"]

            wb = Workbook()
            ws = wb.active
            ws.title = titulo[:31]  # Excel limita a 31 chars

            # --- fila de título ---
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=len(columnas))
            titulo_cell = ws.cell(row=1, column=1, value=titulo)
            titulo_cell.font = Font(bold=True, size=14, color="FFFFFF")
            titulo_cell.fill = PatternFill("solid", fgColor="1a365d")
            titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # --- fila de fecha generación ---
            ws.merge_cells(start_row=2, start_column=1,
                           end_row=2, end_column=len(columnas))
            fecha_cell = ws.cell(
                row=2, column=1,
                value=f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            )
            fecha_cell.font = Font(italic=True, size=10, color="718096")
            fecha_cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[2].height = 18

            # --- cabeceras de columnas ---
            header_fill = PatternFill("solid", fgColor="4a90d9")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            thin = Side(style="thin", color="CBD5E0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col_idx, cabecera in enumerate(cabeceras, start=1):
                cell = ws.cell(row=3, column=col_idx, value=cabecera)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[3].height = 22

            # --- filas de datos ---
            fill_par = PatternFill("solid", fgColor="EBF4FF")
            fill_impar = PatternFill("solid", fgColor="FFFFFF")

            for row_idx, fila in enumerate(datos, start=4):
                fill = fill_par if row_idx % 2 == 0 else fill_impar
                for col_idx, col_key in enumerate(columnas, start=1):
                    valor = fila.get(col_key, "")
                    if valor is None:
                        valor = ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")

            # --- auto-ancho de columnas ---
            for col_idx, col_key in enumerate(columnas, start=1):
                max_len = len(str(cabeceras[col_idx - 1]))
                for fila in datos:
                    val = str(fila.get(col_key, "") or "")
                    max_len = max(max_len, len(val))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

            # --- congelar cabeceras ---
            ws.freeze_panes = "A4"

            wb.save(ruta)
            logger.info(f"Reporte '{titulo}' exportado a Excel: {ruta}")
            return True, None

        except Exception as e:
            AppLogger.log_exception(logger, "Error al exportar a Excel")
            return None, sanitize_error_message(e)

    # ------------------------------------------------------------------ #
    # Exportar a PDF                                                       #
    # ------------------------------------------------------------------ #

    def exportar_pdf(self, clave_reporte, datos, ruta):
        """
        Exporta los datos de un reporte a un archivo PDF.

        Args:
            clave_reporte: clave del dict REPORTES
            datos: lista de dicts
            ruta: ruta completa del archivo destino

        Returns:
            (True, None) si éxito, (None, mensaje_error) si falla
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT

            cfg = REPORTES[clave_reporte]
            columnas = cfg["columnas"]
            cabeceras = cfg["cabeceras"]
            titulo = cfg["titulo"]

            # usar landscape si hay más de 5 columnas
            pagesize = landscape(A4) if len(columnas) > 5 else A4
            doc = SimpleDocTemplate(
                ruta,
                pagesize=pagesize,
                leftMargin=1.5 * cm,
                rightMargin=1.5 * cm,
                topMargin=1.5 * cm,
                bottomMargin=1.5 * cm,
            )

            styles = getSampleStyleSheet()
            story = []

            # --- título ---
            style_titulo = ParagraphStyle(
                "titulo",
                parent=styles["Title"],
                fontSize=16,
                textColor=colors.HexColor("#1a365d"),
                spaceAfter=4,
                alignment=TA_CENTER,
            )
            story.append(Paragraph(titulo, style_titulo))

            # --- subtítulo con fecha ---
            style_sub = ParagraphStyle(
                "subtitulo",
                parent=styles["Normal"],
                fontSize=9,
                textColor=colors.HexColor("#718096"),
                spaceAfter=12,
                alignment=TA_CENTER,
            )
            story.append(Paragraph(
                f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}",
                style_sub
            ))

            story.append(Spacer(1, 0.3 * cm))

            # --- tabla de datos ---
            color_header     = colors.HexColor("#4a90d9")
            color_fila_par   = colors.HexColor("#EBF4FF")
            color_fila_impar = colors.white
            color_borde      = colors.HexColor("#CBD5E0")
            color_texto      = colors.HexColor("#2d3748")

            # estilos de Paragraph para soporte de word-wrap en celdas
            style_hdr = ParagraphStyle(
                "th", fontSize=9, leading=11,
                textColor=colors.white, fontName="Helvetica-Bold",
                alignment=TA_CENTER, wordWrap="LTR",
            )
            style_cel = ParagraphStyle(
                "td", fontSize=8, leading=10,
                textColor=color_texto, fontName="Helvetica",
                alignment=TA_LEFT, wordWrap="LTR",
            )

            # construir filas usando Paragraph para habilitar word-wrap
            filas_tabla = [[Paragraph(h, style_hdr) for h in cabeceras]]
            for fila in datos:
                filas_tabla.append([
                    Paragraph(str(fila.get(col, "") or ""), style_cel)
                    for col in columnas
                ])

            # --- anchos proporcionales según longitud máxima de contenido ---
            page_w = pagesize[0] - 3 * cm  # ancho útil (márgenes 1.5 cm c/lado)
            MIN_COL = 1.5 * cm

            max_lens = []
            for col_key, cabecera in zip(columnas, cabeceras):
                max_len = len(str(cabecera))
                for fila in datos[:100]:          # muestra máximo 100 filas
                    val = str(fila.get(col_key, "") or "")
                    max_len = max(max_len, len(val))
                max_lens.append(max(max_len, 4))  # mínimo 4 caracteres

            total_chars = sum(max_lens)
            col_widths = [
                max((length / total_chars) * page_w, MIN_COL)
                for length in max_lens
            ]
            # escalar para ocupar exactamente el ancho de página
            scale = page_w / sum(col_widths)
            col_widths = [w * scale for w in col_widths]

            tabla = Table(filas_tabla, colWidths=col_widths, repeatRows=1)

            # estilo base
            ts = TableStyle([
                # cabecera
                ("BACKGROUND",  (0, 0), (-1, 0), color_header),
                ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
                ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
                ("LINEBELOW",   (0, 0), (-1, 0), 1.5, color_header),
                # filas alternadas
                ("ROWBACKGROUND", (0, 1), (-1, -1), [color_fila_impar, color_fila_par]),
                # bordes
                ("GRID",        (0, 0), (-1, -1), 0.5, color_borde),
                # padding uniforme
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 6),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ])
            tabla.setStyle(ts)
            story.append(tabla)

            # --- total de registros ---
            story.append(Spacer(1, 0.4 * cm))
            story.append(Paragraph(
                f"Total de registros: {len(datos)}",
                ParagraphStyle("pie", parent=styles["Normal"],
                               fontSize=8, textColor=colors.HexColor("#718096"),
                               alignment=TA_LEFT)
            ))

            doc.build(story)
            logger.info(f"Reporte '{titulo}' exportado a PDF: {ruta}")
            return True, None

        except Exception as e:
            AppLogger.log_exception(logger, "Error al exportar a PDF")
            return None, sanitize_error_message(e)

    @property
    def info_reportes(self):
        return REPORTES
